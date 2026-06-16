#!/usr/bin/env python3
"""
SoS Notebook Automated Tester
Tests all `sos run` cells in xQTL-protocol notebooks listed in
notebook_test_report.xlsx and writes results to status_auto / auto_notes columns.

Usage:
  python3 sos_notebook_tester.py                          # test all untested notebooks
  python3 sos_notebook_tester.py --notebook <rel_path>   # test one notebook
  python3 sos_notebook_tester.py --force-retest           # re-test even if already filled
"""

import argparse
import json
import os
import re
import subprocess
import sys
import time
from pathlib import Path

import openpyxl

# ── Paths ─────────────────────────────────────────────────────────────────────
PROTOCOL_ROOT = Path("/mnt/lustre/lab/gwang/home/rl3328/evo2/xqtl-protocol")
SOS_DIR       = PROTOCOL_ROOT / "code" / "SoS"
EXCEL_PATH    = SOS_DIR / "notebook_test_report.xlsx"

# Per-command timeout (seconds)
CMD_TIMEOUT = 900   # 15 min

# Max character length for Excel cell notes (openpyxl limit is 32767, but keep it readable)
MAX_NOTE_LEN = 500

# ── ANSI stripping ─────────────────────────────────────────────────────────────
_ANSI_RE = re.compile(r"\x1b\[[0-9;]*[mGKHF]|\x1b\].*?\x07|\x1b[@-_][0-?]*[ -/]*[@-~]")

def strip_ansi(text: str) -> str:
    """Remove all ANSI escape sequences (color codes etc.) from text."""
    return _ANSI_RE.sub("", text)

def clean_note(text: str) -> str:
    """Strip ANSI codes, collapse whitespace, truncate to MAX_NOTE_LEN."""
    t = strip_ansi(text)
    t = re.sub(r"\s+", " ", t).strip()
    if len(t) > MAX_NOTE_LEN:
        t = t[:MAX_NOTE_LEN - 3] + "..."
    return t


# ── Notebook parsing ───────────────────────────────────────────────────────────

def extract_sos_commands(nb_path: Path) -> list[tuple[int, str]]:
    """
    Return [(cell_idx, raw_source), ...] for every testable sos-run cell.

    Skipped:
      - Non-code cells
      - Cells starting with `!` (Jupyter shell magic)
      - Cells whose first sos run token is followed by -h
    """
    with open(nb_path) as f:
        nb = json.load(f)

    cmds = []
    for idx, cell in enumerate(nb["cells"]):
        if cell["cell_type"] != "code":
            continue
        src = "".join(cell["source"]).strip()
        if not src:
            continue

        # Skip Jupyter magic shell runs
        if src.startswith("!"):
            continue

        # Skip help commands: first `sos run` line ends in -h
        first_sos_line = next(
            (l.strip() for l in src.splitlines() if "sos run" in l), ""
        )
        if re.search(r"\bsos run\b.*\s-h\s*$", first_sos_line):
            continue

        # Accept cells whose first non-comment, non-blank line is `sos run`
        # (handles "# Step N — …\nsos run …" patterns), bash for-loop wrappers,
        # and plain sos run cells.
        first_code_line = next(
            (l.strip() for l in src.splitlines()
             if l.strip() and not l.strip().startswith("#")),
            ""
        )
        if (
            first_code_line.startswith("sos run")
            or (src.startswith("for ") and "sos run" in src)
        ):
            cmds.append((idx, src))

    return cmds


def inject_force(src: str) -> str:
    """Append `-s force` to every `sos run` line that doesn't already have it."""
    out_lines = []
    for line in src.splitlines():
        if re.match(r"\s*sos run\b", line) and "-s force" not in line:
            stripped = line.rstrip()
            if stripped.endswith("\\"):
                stripped = stripped[:-1].rstrip() + " -s force \\"
            else:
                stripped = stripped + " -s force"
            out_lines.append(stripped)
        else:
            out_lines.append(line)
    return "\n".join(out_lines)


# ── Input validation ───────────────────────────────────────────────────────────

_FILE_ARG_RE = re.compile(r"--[\w-]*[Ff]ile\s+(\S+)")
_INPUT_ARG_RE = re.compile(r"--inputs?\s+(\S+)")

def check_inputs(src: str) -> tuple[bool, str | None]:
    """
    Verify that every file referenced by --*File / --input arguments exists.
    Skips paths containing shell variables ($ characters).
    Returns (True, None) if all exist, (False, missing_rel_path) otherwise.
    """
    candidates = _FILE_ARG_RE.findall(src) + _INPUT_ARG_RE.findall(src)
    for rel in candidates:
        # Skip shell variable expansions: $var, ${var}, $_, backtick substitutions
        if "$" in rel or "`" in rel:
            continue
        # Only check paths that look like output/ or input/ relative paths
        if not (rel.startswith("output/") or rel.startswith("input/")):
            continue
        full = PROTOCOL_ROOT / rel
        if not full.exists():
            return False, rel
    return True, None


# ── Command execution ──────────────────────────────────────────────────────────

def run_command(src: str, timeout: int = CMD_TIMEOUT) -> tuple[int, str, float]:
    """
    Execute src as a bash script from PROTOCOL_ROOT.
    Returns (returncode, error_snippet, elapsed_seconds).
    """
    full_script = "set -o pipefail\n" + src

    t0 = time.time()
    try:
        result = subprocess.run(
            full_script,
            shell=True,
            executable="/bin/bash",
            cwd=PROTOCOL_ROOT,
            capture_output=True,
            text=True,
            timeout=timeout,
        )
        elapsed = time.time() - t0

        # Build a compact, clean error snippet
        lines = strip_ansi(result.stderr).splitlines()
        err_lines = [
            l for l in lines
            if any(kw in l for kw in ("ERROR", "Error", "error", "Traceback", "FAILED", "failed"))
        ]
        snippet = (err_lines[-1] if err_lines else (lines[-1] if lines else "")).strip()
        return result.returncode, snippet[:250], elapsed

    except subprocess.TimeoutExpired:
        return -1, f"TIMEOUT after {timeout}s", float(timeout)


# ── Notebook tester ────────────────────────────────────────────────────────────

def test_notebook(rel_path: str) -> tuple[str, str, float]:
    """
    Run all testable sos run cells in the notebook at SOS_DIR/rel_path.

    Returns (status, notes, total_elapsed_seconds).
    status ∈ {'PASS', 'FAIL', 'SKIP', 'ERROR'}

    Logic:
    - If no testable cells or notebook missing → SKIP
    - First cell: missing input → SKIP (can't even start)
    - Subsequent cells: missing input after prior failures → treated as cascade FAIL
    - Any non-zero exit code → FAIL (all failures collected, run continues)
    - All exit 0 → PASS
    """
    nb_path = SOS_DIR / rel_path

    if not nb_path.exists():
        return "SKIP", f"Notebook file not found", 0.0

    try:
        cmds = extract_sos_commands(nb_path)
    except Exception as exc:
        return "ERROR", f"Failed to parse notebook: {exc}", 0.0

    if not cmds:
        return "SKIP", "No testable sos run cells found", 0.0

    total_elapsed = 0.0
    fail_notes: list[str] = []
    is_first_cell = True

    for cell_idx, raw_src in cmds:
        src = inject_force(raw_src)

        # Input check: SKIP only if first cell is missing input (can't start at all).
        # For subsequent cells, a missing input is a cascade from prior failure → FAIL.
        ok, missing = check_inputs(src)
        if not ok:
            if is_first_cell:
                return "SKIP", f"Missing input file: {missing}", 0.0
            else:
                fail_notes.append(f"Cell {cell_idx}: missing input {missing} (cascade)")
                break   # No point continuing if intermediate output is absent

        is_first_cell = False

        print(f"    Cell {cell_idx}: running ...", flush=True)
        rc, snippet, elapsed = run_command(src)
        total_elapsed += elapsed

        if rc != 0:
            fail_notes.append(f"Cell {cell_idx} (exit {rc}): {snippet}")
            print(f"    Cell {cell_idx}: FAILED in {elapsed:.1f}s — {snippet[:80]}", flush=True)
            # Continue to next cell so we surface all failures in one run
        else:
            print(f"    Cell {cell_idx}: OK in {elapsed:.1f}s", flush=True)

    if fail_notes:
        return "FAIL", clean_note("; ".join(fail_notes)), total_elapsed

    return "PASS", "All sos run cells completed (exit 0)", total_elapsed


# ── Excel helpers ──────────────────────────────────────────────────────────────

HEADER_ROW_MARKER = "#"
STATUS_AUTO_HEADER = "status_auto"
AUTO_NOTES_HEADER  = "auto_notes"


def load_workbook_with_cols(path: Path):
    """Load workbook and return (wb, ws, header_row_num, col_status_auto, col_auto_notes)."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    header_row_num = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row[0] == HEADER_ROW_MARKER:
            header_row_num = i
            break
    if header_row_num is None:
        raise ValueError(f"Could not find header row in {path}")

    col_map: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 5):
        val = ws.cell(header_row_num, col_idx).value
        if val is not None:
            col_map[val] = col_idx

    if STATUS_AUTO_HEADER not in col_map:
        new_col = max(col_map.values()) + 1
        ws.cell(header_row_num, new_col).value = STATUS_AUTO_HEADER
        col_map[STATUS_AUTO_HEADER] = new_col

    if AUTO_NOTES_HEADER not in col_map:
        new_col = col_map[STATUS_AUTO_HEADER] + 1
        ws.cell(header_row_num, new_col).value = AUTO_NOTES_HEADER
        col_map[AUTO_NOTES_HEADER] = new_col

    return wb, ws, header_row_num, col_map[STATUS_AUTO_HEADER], col_map[AUTO_NOTES_HEADER]


def iter_notebook_rows(ws, header_row_num: int):
    """Yield (row_num, row_id, rel_path) for each data row before the legend."""
    for row_num in range(header_row_num + 1, ws.max_row + 1):
        row_id   = ws.cell(row_num, 1).value
        rel_path = ws.cell(row_num, 3).value

        if row_id == "Status Legend":
            break
        if row_id is None or rel_path is None:
            continue
        if not str(row_id).isdigit() and not isinstance(row_id, int):
            continue

        yield row_num, row_id, str(rel_path)


def write_result(wb, ws, row_num, col_auto, col_notes, col_time,
                 status, notes, elapsed):
    """Write a single result row and save."""
    ws.cell(row_num, col_auto).value  = status
    ws.cell(row_num, col_notes).value = clean_note(notes)
    if ws.cell(row_num, col_time).value is None and elapsed > 0:
        ws.cell(row_num, col_time).value = f"{elapsed:.0f}s"
    wb.save(EXCEL_PATH)


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="SoS notebook automated tester")
    parser.add_argument("--notebook",     help="Relative path of a single notebook to test")
    parser.add_argument("--force-retest", action="store_true",
                        help="Re-test even if status_auto is already filled")
    args = parser.parse_args()

    wb, ws, header_row, col_auto, col_notes = load_workbook_with_cols(EXCEL_PATH)
    col_time = 6

    rows = list(iter_notebook_rows(ws, header_row))

    if args.notebook:
        rows = [(r, i, p) for r, i, p in rows if p == args.notebook]
        if not rows:
            print(f"ERROR: notebook '{args.notebook}' not found in Excel.", file=sys.stderr)
            sys.exit(1)

    tested = skipped = failed = passed = 0

    for row_num, row_id, rel_path in rows:
        existing = ws.cell(row_num, col_auto).value
        if existing and not args.force_retest:
            print(f"[{row_id}] {rel_path} — already {existing}, skipping")
            continue

        print(f"\n[{row_id}] Testing: {rel_path}", flush=True)
        tested += 1

        try:
            status, notes, elapsed = test_notebook(rel_path)
        except Exception as exc:
            status, notes, elapsed = "ERROR", str(exc)[:300], 0.0

        print(f"  → {status}: {clean_note(notes)[:100]} ({elapsed:.0f}s)", flush=True)

        write_result(wb, ws, row_num, col_auto, col_notes, col_time,
                     status, notes, elapsed)

        if status == "PASS":   passed  += 1
        elif status == "FAIL": failed  += 1
        else:                  skipped += 1

    print(f"\n{'='*60}")
    print(f"Done. Tested={tested}  PASS={passed}  FAIL={failed}  SKIP/OTHER={skipped}")
    print(f"Results saved to {EXCEL_PATH}")


if __name__ == "__main__":
    main()
